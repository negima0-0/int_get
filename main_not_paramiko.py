from netmiko import ConnectHandler
import json
import datetime
import csv
import openpyxl
from openpyxl import load_workbook

def extract_interface_data(json_output):
    interface_data = []
    interfaces = json.loads(json_output)['configuration']['interfaces']['interface']
    for interface in interfaces:
        name = interface['name']
        multicast_in = interface['traffic-statistics']['multicast-packets']['input']
        multicast_out = interface['traffic-statistics']['multicast-packets']['output']
        broadcast_in = interface['traffic-statistics']['broadcast-packets']['input']
        broadcast_out = interface['traffic-statistics']['broadcast-packets']['output']
        interface_data.append((name, multicast_in, multicast_out, broadcast_in, broadcast_out))
    return interface_data

def write_to_excel(data, filename, hostname):
    try:
        workbook = load_workbook(filename)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(['Date', 'Hostname', 'Multicast (RX)', 'Multicast (TX)', 'Broadcast (RX)', 'Broadcast (TX)'])

    for row, (name, multicast_in, multicast_out, broadcast_in, broadcast_out) in enumerate(data, start=2):
        worksheet.append([datetime.date.today().strftime('%Y-%m-%d'), hostname, multicast_in, multicast_out, broadcast_in, broadcast_out])

    workbook.save(filename)

def main():
    # CSVファイルからホストの情報を読み取る
    hosts = []
    with open('hosts.csv', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            hostname = row.get('hostname')
            ip = row.get('ip')
            if not hostname and not ip:
                continue

            hosts.append({'hostname': hostname, 'ip': ip})

    # 踏み台経由で各ホストにアクセスし、データを取得してExcelに書き込む
    for host_info in hosts:
        hostname = host_info['hostname']
        ip = host_info['ip']
        print(f"Processing host: {hostname or ip}")
        ssh_session = ConnectHandler(
            device_type='juniper_junos',
            host=ip,
            username='your_username',  # ここにユーザ名を入力
            password='your_password',  # ここにパスワードを入力
            use_paramiko=False
        )
        output = ssh_session.send_command('show interface extensive | display json')
        ssh_session.disconnect()

        interface_data = extract_interface_data(output)
        write_to_excel(interface_data, 'interface_stats.xlsx', hostname or ip)
        print(f"Completed processing host: {hostname or ip}")

if __name__ == "__main__":
    main()
