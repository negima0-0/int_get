from openpyxl import Workbook, load_workbook
from datetime import datetime

# エクセルファイルのパス
file_path = "data.xlsx"

# データ
data1 = [1, 2, 3, 4, 5]
data2 = ['a', 'b', 'c', 'd', 'e']
data3 = [True, False, True, False, True]
data4 = [3.14, 2.718, 1.618, 0.577, 1.414]
data5 = ['apple', 'banana', 'orange', 'grape', 'watermelon']

# 日時
now = datetime.now()

# 新しいワークブックを作成するか、既存のワークブックを読み込む
try:
    wb = load_workbook(file_path)
except FileNotFoundError:
    wb = Workbook()

# シートが存在しない場合は作成し、存在する場合は取得
if "Sheet1" not in wb.sheetnames:
    sheet = wb.create_sheet(title="Sheet1")
else:
    sheet = wb["Sheet1"]

# データをエクセルファイルに書き込む
row = [now] + data1 + data2 + data3 + data4 + data5
sheet.append(row)

# エクセルファイルを保存
wb.save(file_path)
